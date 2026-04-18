"""
Re-imports all products and providers from XLS files into the database.
- Patara_catalog.xlsx  → items for provider "Patara"
- ספקים.xlsx           → provider records (iron / concrete / main suppliers)
"""
import openpyxl
import psycopg2
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

DB = dict(host='localhost', port=5432, dbname='builtin', user='postgres', password='')
CATALOG   = 'BE/src/resources/Patara_catalog.xlsx'
SUPPLIERS = 'BE/src/main/resources/ספקים.xlsx'


# ── unit detection from a single text fragment ───────────────────────────────
def detect_unit(text: str) -> str:
    if not text:
        return 'UNIT'
    d = str(text).lower()
    t = str(text)
    if 'kg' in d or 'ק"ג' in t or 'קג' in t:           return 'KG'
    if 'gram' in d or 'גרם' in t:                        return 'GRAM'
    if 'liter' in d or 'ליטר' in t:                      return 'LITER'
    if 'ml' in d or "מ'ל" in t or 'מל' in t or 'cc' in d: return 'MILLILITER'
    if 'מ"מ' in t or ' mm' in d or d.endswith('mm'):     return 'MILLIMETER'
    if ' cm' in d or d.endswith('cm') or 'ס"מ' in t:     return 'CENTIMETER'
    if 'מ"ר' in t or 'm2' in d or 'sqm' in d:           return 'SQUARE_METER'
    if 'meter' in d or 'מטר' in t:                       return 'METER'
    if 'pack' in d or 'חבילה' in t:                      return 'PACK'
    if 'box' in d or 'קופסה' in t or 'ארגז' in t:        return 'BOX'
    if 'dozen' in d or 'תריסר' in t:                     return 'DOZEN'
    return 'UNIT'


# ── parse disc column ─────────────────────────────────────────────────────────
# Format: [pack/unit] | [stock qty] | [measurement/unit] | [notes ...]
def parse_disc(disc: str):
    """Returns (quantity: int, unit: str) extracted from the disc field."""
    if not disc:
        return 1, 'UNIT'

    parts = [p.strip() for p in str(disc).split('|')]
    quantity = 1
    unit = 'UNIT'

    for i, part in enumerate(parts):
        # Detect unit keyword anywhere in the parts
        u = detect_unit(part)
        if u != 'UNIT':
            unit = u

    # Stock quantity is the SECOND part (index 1) if it is a pure number
    if len(parts) >= 2:
        try:
            val = float(parts[1].replace(',', '.'))
            if val > 0:
                quantity = int(val)
        except (ValueError, TypeError):
            pass

    # If no stock in part 2, fall back to first part if it's a pure number
    if quantity == 1 and len(parts) >= 1:
        try:
            val = float(parts[0].replace(',', '.'))
            if val > 0:
                quantity = int(val)
        except (ValueError, TypeError):
            pass

    return max(quantity, 1), unit


# ── helpers ───────────────────────────────────────────────────────────────────
def upsert_provider(cur, name, phone=None, location=None):
    email = name.lower().replace(' ', '_').replace('.', '') + '@builtin.com'
    cur.execute('SELECT id FROM providers WHERE name = %s', (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        'INSERT INTO providers (name, phone, location, email) VALUES (%s,%s,%s,%s) RETURNING id',
        (name, phone, location, email)
    )
    return cur.fetchone()[0]


def main():
    conn = psycopg2.connect(**DB)
    cur  = conn.cursor()

    # ── 1. wipe old data ──────────────────────────────────────────────────────
    print('Clearing old items and providers…')
    cur.execute('DELETE FROM item_photos')
    cur.execute('DELETE FROM items')
    cur.execute('DELETE FROM users')
    cur.execute('DELETE FROM providers')
    conn.commit()

    # ── 2. providers from ספקים.xlsx ─────────────────────────────────────────
    print('Importing providers from ספקים.xlsx…')
    wb_sup = openpyxl.load_workbook(SUPPLIERS, read_only=True)
    provider_count = 0
    for sheet_name in wb_sup.sheetnames:
        ws = wb_sup[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:          # skip header
            name = row[0]
            phone = row[3] if len(row) > 3 else None
            if not name or str(name).strip() == '':
                continue
            name = str(name).strip()
            phone = str(phone).strip() if phone else None
            upsert_provider(cur, name, phone=phone)
            provider_count += 1
    conn.commit()
    print(f'  Created {provider_count} supplier records')

    # ── 3. Patara provider + items from Patara_catalog.xlsx ──────────────────
    print('Importing Patara catalog…')
    patara_id = upsert_provider(cur, 'Patara', location='Israel')
    conn.commit()
    print(f'  Patara provider id = {patara_id}')

    wb_cat = openpyxl.load_workbook(CATALOG, read_only=True)
    ws     = wb_cat['מוצרים']
    rows   = list(ws.iter_rows(values_only=True))

    # locate columns by header
    header = [str(c).strip().lower() if c else '' for c in rows[0]]
    try:
        idx_serial = header.index('serial')
        idx_name   = header.index('name')
        idx_disc   = header.index('disc')
        idx_price  = header.index('price')
    except ValueError:
        print('  Header:', header)
        raise RuntimeError('Could not find expected columns (serial, name, disc, price)')

    inserted = skipped = errors = 0
    for row in rows[1:]:
        if not row or row[idx_name] is None:
            skipped += 1
            continue

        serial = str(row[idx_serial]).strip()[:255] if row[idx_serial] is not None else None
        name   = str(row[idx_name]).strip()[:255]
        disc   = str(row[idx_disc]).strip()[:255] if row[idx_disc] is not None else None
        try:
            price = float(row[idx_price]) if row[idx_price] is not None else 0.0
        except (TypeError, ValueError):
            price = 0.0

        quantity, unit = parse_disc(disc)

        try:
            cur.execute(
                '''INSERT INTO items (name, item_type, serial_number, price, quantity, unit, provider_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (serial_number) DO NOTHING''',
                (name, disc, serial, price, quantity, unit, patara_id)
            )
            inserted += (1 if cur.rowcount else 0)
            skipped  += (0 if cur.rowcount else 1)
        except Exception as e:
            print(f'  Error serial={serial}: {e}')
            conn.rollback()
            errors += 1

    conn.commit()
    print(f'  Items inserted: {inserted}, skipped: {skipped}, errors: {errors}')

    # ── 4. create Patara user account ─────────────────────────────────────────
    print('Creating Patara provider user…')
    cur.execute(
        '''INSERT INTO users (first_name, last_name, email, user_type, provider_id)
           VALUES (%s,%s,%s,%s,%s)
           ON CONFLICT (email) DO NOTHING''',
        ('Patara', 'Admin', 'patara@builtin.com', 'PROVIDER', patara_id)
    )
    conn.commit()
    print('  Done.')

    cur.execute('SELECT COUNT(*) FROM items')
    print(f'\nTotal items in DB: {cur.fetchone()[0]}')
    cur.execute('SELECT COUNT(*) FROM providers')
    print(f'Total providers in DB: {cur.fetchone()[0]}')

    conn.close()
    print('\nAll done!')


if __name__ == '__main__':
    main()
